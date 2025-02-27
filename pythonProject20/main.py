import openpyxl
import re
from openpyxl.utils import get_column_letter, column_index_from_string
import gspread
   # search_text = input('Что ищем')
search_text = str("Яков")
print('Ищем:', search_text)
wb = openpyxl.load_workbook("Data.xlsx")  # Грузим наш прайс-лист
sheets_list = wb.sheetnames  # Получаем список всех листов в файле
sheet_active = wb[sheets_list[0]]  # Начинаем работать с самым первым
row_max = sheet_active.max_row  # Получаем количество столбцов
column_max = sheet_active.max_column  # Получаем количество строк
row_min = 1 #Переменная, отвечающая за номер строки
column_min = 1 #Переменная, отвечающая за номер столбца
while column_min <= column_max:
    row_min_min = row_min
    row_max_max = row_max
    while row_min_min <= row_max_max:
        row_min_min = str(row_min_min)
        word_column = get_column_letter(column_min)
        word_column = str(word_column)
        word_cell = word_column + row_min_min
        data_from_cell = sheet_active[word_cell].value
        data_from_cell = str(data_from_cell)
        regular = search_text
        result = re.findall(regular, data_from_cell)
        ws = wb.active
        if len(result) > 0:
            print('Нашли в ячейке:', word_cell, column_index_from_string(word_column))
        row_min_min = int(row_min_min)
        row_min_min = row_min_min + 1
    column_min = column_min + 1